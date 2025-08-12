import os
import sys
from datetime import datetime
from typing import List, Dict, Any, Tuple

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape


DEFAULT_EXCEL_PATH = "/workspace/Planilha SOSrim.xlsx"
OUTPUT_DIR = "/workspace/site"
TEMPLATE_DIR = "/workspace/site"
TEMPLATE_NAME = "template.html"


def normalize_header(cell_value: Any, fallback_index: int) -> str:
    if cell_value is None or str(cell_value).strip() == "":
        return f"Coluna {fallback_index+1}"
    return str(cell_value).strip()


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        text = ("%f" % value).rstrip("0").rstrip(".") if isinstance(value, float) else str(value)
        return text
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return str(value)


def count_non_empty(row: Tuple[Any, ...]) -> int:
    return sum(1 for c in row if c is not None and str(c).strip() != "")


def trim_trailing_empty(values: List[Any]) -> List[Any]:
    end = len(values) - 1
    while end >= 0 and (values[end] is None or str(values[end]).strip() == ""):
        end -= 1
    return values[: end + 1]


def read_workbook(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        # Look at the first up to 50 rows to identify the header row as the one with most non-empty cells
        header_candidate_idx = None
        header_candidate_values: Tuple[Any, ...] | None = None
        max_filled = 0
        lookahead_rows: List[Tuple[Any, ...]] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx > 50:
                break
            lookahead_rows.append(row)
            filled = count_non_empty(row)
            if filled > max_filled:
                max_filled = filled
                header_candidate_idx = row_idx
                header_candidate_values = row
        if header_candidate_idx is None or header_candidate_values is None or max_filled < 2:
            # Skip sheets that appear empty or without a usable header
            continue

        headers_raw = trim_trailing_empty(list(header_candidate_values))
        headers = [normalize_header(v, idx) for idx, v in enumerate(headers_raw)]
        num_cols = len(headers)
        if num_cols == 0:
            continue

        data_rows: List[List[str]] = []
        # Start reading from the row after the header candidate until the end
        for row in ws.iter_rows(min_row=header_candidate_idx + 1, values_only=True):
            # Consider only the first num_cols cells
            limited = list(row[:num_cols]) if row is not None else []
            # Skip fully empty rows
            if count_non_empty(tuple(limited)) == 0:
                continue
            normalized = [normalize_cell(cell) for cell in limited]
            if len(normalized) < num_cols:
                normalized += [""] * (num_cols - len(normalized))
            elif len(normalized) > num_cols:
                normalized = normalized[:num_cols]
            data_rows.append(normalized)

        sheets.append({
            "name": ws.title,
            "headers": headers,
            "rows": data_rows,
        })
    return sheets


def render_site(sheets: List[Dict[str, Any]], source_name: str) -> str:
    env = Environment(
        loader=FileSystemLoader(TEMPLATE_DIR),
        autoescape=select_autoescape(['html', 'xml'])
    )
    template = env.get_template(TEMPLATE_NAME)
    html = template.render(
        title="Planilha SOSrim — Consulta",
        generated_at=datetime.now().strftime("%d/%m/%Y %H:%M"),
        sheet_count=len(sheets),
        sheets=sheets,
        source_name=os.path.basename(source_name),
    )
    return html


def main() -> None:
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {excel_path}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    sheets = read_workbook(excel_path)
    html = render_site(sheets, excel_path)
    output_path = os.path.join(OUTPUT_DIR, "index.html")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Site gerado em: {output_path}")


if __name__ == "__main__":
    main()