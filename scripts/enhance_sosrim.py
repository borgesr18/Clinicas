#!/usr/bin/env python3
import sys
from pathlib import Path
from typing import Optional, Tuple
import datetime as _dt

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


INPUT_PATH = Path('/workspace/Planilha SOSrim.xlsx')
OUTPUT_PATH = Path('/workspace/Planilha SOSrim (Melhorada).xlsx')
DASHBOARD_SHEET_NAME = 'Dashboard_Auto'


def is_cell_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    return False


def detect_used_bounds(ws) -> Tuple[int, int, int, int]:
    """Return (min_row, min_col, max_row, max_col) for non-empty cells, fallback to ws.max_* if needed."""
    max_row = 0
    max_col = 0
    min_row: Optional[int] = None
    min_col: Optional[int] = None

    # Iterate within reported bounds, but skip completely empty rows/cols
    for row in ws.iter_rows(values_only=True):
        # row is a tuple of values
        if any(not is_cell_empty(v) for v in row):
            r_idx = row[0:0].__len__()  # dummy to keep linter happy
            # We don't have index here; we will re-iterate with enumerate to get positions
            min_row = 1
            break

    # More reliable approach: scan by coordinates to detect min/max
    for r in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(not is_cell_empty(v) for v in row_values):
            if min_row is None:
                min_row = r
            max_row = r
            for c, v in enumerate(row_values, start=1):
                if not is_cell_empty(v):
                    if min_col is None or c < min_col:
                        min_col = c
                    if c > max_col:
                        max_col = c

    if min_row is None:
        # Empty sheet fallback
        return (1, 1, 1, 1)

    return (min_row, min_col or 1, max_row or min_row, max_col or 1)


def style_header_row(ws, header_row: int, min_col: int, max_col: int) -> None:
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
    ws.row_dimensions[header_row].height = 22


def adjust_column_widths(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    max_width_by_col = {}
    max_sample_rows = min(max_row, min_row + 500)  # sample at most 500 rows for performance

    for col in range(min_col, max_col + 1):
        max_len = 0
        # include header
        header_value = ws.cell(row=min_row, column=col).value
        if header_value is not None:
            max_len = max(max_len, len(str(header_value)))
        for row in range(min_row + 1, max_sample_rows + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            text = value if isinstance(value, str) else str(value)
            if len(text) > max_len:
                max_len = len(text)
        # heuristic conversion char count to Excel width
        width = min(max(10, int(max_len * 1.1) + 2), 60)
        max_width_by_col[col] = width

    for col, width in max_width_by_col.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_autofilter(ws, header_row: int, min_col: int, max_col: int, max_row: int) -> None:
    if max_row <= header_row:
        return
    ws.auto_filter.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{max_row}"


def add_table(ws, header_row: int, min_col: int, max_col: int, max_row: int) -> None:
    if max_row <= header_row:
        return
    # openpyxl tables require contiguous data and unique header names
    # Build table name from sheet name and ensure it's unique and valid
    base_name = ''.join(ch for ch in ws.title if ch.isalnum()) or 'Tabela'
    table_name = base_name[:20] + 'Table'
    # Ensure uniqueness among existing tables
    existing = {t.name for t in ws._tables}
    suffix = 1
    unique_name = table_name
    while unique_name in existing:
        unique_name = f"{table_name}_{suffix}"
        suffix += 1

    ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{max_row}"
    try:
        table = Table(displayName=unique_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    except Exception:
        # If headers are invalid/duplicate, skip silently to avoid breaking
        pass


def apply_duplicate_highlight(ws, header_row: int, min_col: int, max_col: int, max_row: int) -> None:
    if max_row <= header_row:
        return
    fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    font = Font(color='9C0006')
    dxf = DifferentialStyle(fill=fill, font=font)
    for col in range(min_col, max_col + 1):
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{header_row+1}:{col_letter}{max_row}"
        rule = Rule(type='duplicateValues', dxf=dxf, stopIfTrue=False)
        ws.conditional_formatting.add(cell_range, rule)


def format_dates(ws, header_row: int, min_col: int, max_col: int, max_row: int) -> None:
    # Detect columns with date/datetime values by sampling a subset
    sample_limit = min(max_row, header_row + 200)
    date_columns = set()
    for col in range(min_col, max_col + 1):
        for row in range(header_row + 1, sample_limit + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, (_dt.date, _dt.datetime)):
                date_columns.add(col)
                break
    for col in date_columns:
        for row in range(header_row + 1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (_dt.date, _dt.datetime)):
                cell.number_format = 'DD/MM/YYYY'


def build_dashboard(wb) -> None:
    # Ensure unique dashboard name
    name = DASHBOARD_SHEET_NAME
    existing_titles = {ws.title for ws in wb.worksheets}
    if name in existing_titles:
        idx = 2
        while f"{name}_{idx}" in existing_titles:
            idx += 1
        name = f"{name}_{idx}"
    ws = wb.create_sheet(title=name)

    ws['A1'] = 'Aba'
    ws['B1'] = 'Linhas'
    ws['C1'] = 'Colunas'
    ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)

    data = []
    for sh in wb.worksheets:
        if sh.title == ws.title:
            continue
        min_row, min_col, max_row, max_col = detect_used_bounds(sh)
        # If the top row seems like a header, exclude it from count of data rows
        data_rows = max(0, max_row - min_row)
        data.append((sh.title, data_rows, max_col - min_col + 1))

    start_row = 2
    for i, (title, n_rows, n_cols) in enumerate(data, start=start_row):
        ws.cell(row=i, column=1, value=title)
        ws.cell(row=i, column=2, value=n_rows)
        ws.cell(row=i, column=3, value=n_cols)

    # Create a simple bar chart of rows per sheet
    if data:
        chart = BarChart()
        chart.title = 'Registros por aba'
        chart.y_axis.title = 'Linhas'
        chart.x_axis.title = 'Aba'
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(data) - 1)
        values = Reference(ws, min_col=2, min_row=1, max_row=start_row + len(data) - 1)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 24
        ws.add_chart(chart, 'E2')

    # Basic width adjustments
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12


def enhance_workbook(input_path: Path, output_path: Path) -> None:
    if not input_path.exists():
        print(f"Arquivo não encontrado: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(input_path)

    for ws in wb.worksheets:
        # Skip dashboard sheets during styling to avoid self-formatting
        if ws.title.startswith('Dashboard'):
            continue
        min_row, min_col, max_row, max_col = detect_used_bounds(ws)
        header_row = min_row

        # Freeze header row
        if max_row > header_row:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=min_col)

        # Style header
        style_header_row(ws, header_row, min_col, max_col)

        # Auto filter
        add_autofilter(ws, header_row, min_col, max_col, max_row)

        # Table styling (best-effort)
        add_table(ws, header_row, min_col, max_col, max_row)

        # Adjust column widths
        adjust_column_widths(ws, min_row, min_col, max_row, max_col)

        # Date formatting
        format_dates(ws, header_row, min_col, max_col, max_row)

        # Duplicate highlighting
        apply_duplicate_highlight(ws, header_row, min_col, max_col, max_row)

    # Add dashboard sheet summarizing the workbook
    build_dashboard(wb)

    # Save improved copy
    wb.save(output_path)
    print(f"Arquivo salvo em: {output_path}")


if __name__ == '__main__':
    enhance_workbook(INPUT_PATH, OUTPUT_PATH)